import openpyxl
import os

'''
openxlsx的基本使用：

    base_path = os.path.dirname(os.getcwd())
    open_excel = openpyxl.load_workbook(base_path+"/Case/imooc.xlsx") #打开文件
    
    sheet_name = open_excel.sheetnames # 拿到表格中所有的sheet
    
    excel_value = open_excel[sheet_name[0]] # 获取第一个sheet1
    
    print(excel_value.cell(2,2).value)  # 获取第二行第二个的内容
    
    print(excel_value.max_row)

'''
base_path = os.path.dirname(os.getcwd()) # basepath是F:\API_Autotest



class HandExcel:

    def load_excel(self):
        '''加载excel,获取整个表格，是一个表格对象'''
        open_excel = openpyxl.load_workbook(base_path + "/Case/test.xlsx")  # 打开文件
        return open_excel

    def get_sheet_data(self,index=None):
        '''加载表格中所有的sheet，并获取具体的某一个sheet'''
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index=0

        data = self.load_excel()[sheet_name[index]] # 获取某一个具体sheet
        return data

    def get_cell_value(self,row,cols):
        '''获取某一个单元格内容,通过行和列去指定哪一个单元格'''

        data = self.get_sheet_data().cell(row, cols).value
        return data

    def get_rows(self):
        '''获取所有的行数'''
        row = self.get_sheet_data().max_row
        return row

    def get_row_value(self,row):
        '''获取某一行的内容'''

        row_list=[]

        # self.get_sheet_data()[row] 获取到的是某一行所有单元格的对象

        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)  # .value 获取到具体的值

        return row_list

    def excel_write_data(self,row,cols,value):
        '''
            作用：将内容写入到excel中，目前用于用例执行结果的回写
            注意：row,cols都是从1开始算起

        '''
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row,cols,value)
        wb.save(base_path + "/Case/test.xlsx")


    def get_columns_value(self,key=None):
        '''
        函数作用：获取某一列的内容，用于case依赖的解决
        现在是获取用例编号
        '''
        columns_list = []

        if key == None:
            key = "A"
        columns_list_data = self.get_sheet_data()[key] # 这个key 就是excel中的A B C D这些列号,现在这样返回的是单元格的对象，需要用.value取值出来

        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list # 结果： ['case编号', 'imooc_001', 'imooc_002', 'imooc_003', 'imooc_004', 'imooc_005', 'imooc_006', 'imooc_007', None, None]

    def get_rows_number(self,case_id):
        '''
            函数作用：获取被依赖得case所在的行号，即：传入case_id，用于根据依赖的case_id获取依赖的case在哪一行
        '''
        number = 1 # 表示行号，第一行是标题，势必不会等于，那就是从第二行开始比较，据此推理，行号要从1开始
        cols_data = self.get_columns_value()

        for col_data in cols_data:
            if case_id == col_data:
                return number
            number += 1
        return number

    def get_excel_data(self):
        '''
            获取excel里面所有行的数据,并放到list里面，供ddt使用

        '''
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_row_value(i+2)) # i从0开始，但是excel中第一行是标题，所以是第二行开始，即 i+2
        return data_list




if __name__=="__main__":
    handle = HandExcel()
    print(handle.get_excel_data())
    print(handle.get_rows())
    print(handle.get_row_value(1))
#     print(handle.get_columns_value("A"))
#     print(handle.get_rows_number("imooc_012"))
#     print(handle.excel_write_data(3,13,"成功")) # 这里的1列完全是按照excel左边的数，2代表第2列
# #
# #     a = os.path.dirname(__file__) # F:/API_Autotest/Util
# #
# #     b=__file__  # F:/API_Autotest/Util/handle_excel.py
# #
# #     c=os.path.dirname(__file__)
# #     base_path = os.path.dirname(os.getcwd())  # basepath是F:\API_Autotest
# #
# #     print("basepath是%s" % base_path)
# #     print("a的路径是%s"%a)
# #     print("b的路径是%s"%b)
#     print("c的路径是%s"%c)

